from datetime import datetime
from functools import wraps

import openpyxl
from flask import (
    Blueprint,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    session,
    url_for,
)

from .models import Answer, Question, QuestionOption, Response, Survey, User, db

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin"):
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated


def operator_required(f):
    """Permite acceso a admin y operadores."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("is_admin") and not session.get("is_operator"):
            return redirect(url_for("auth.login"))
        return f(*args, **kwargs)
    return decorated


# ── Dashboard ──────────────────────────────────────────────────────────────────

@admin_bp.route("/")
@admin_required
def index():
    surveys = Survey.query.order_by(Survey.created_at.desc()).all()
    users_count = User.query.filter_by(is_admin=False).count()
    return render_template("admin/index.html", surveys=surveys, users_count=users_count)


# ── Survey CRUD ────────────────────────────────────────────────────────────────

@admin_bp.route("/survey/new", methods=["GET", "POST"])
@operator_required
def survey_new():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        description = request.form.get("description", "").strip()
        start_str = request.form.get("start_date", "").strip()
        end_str = request.form.get("end_date", "").strip()
        is_active = request.form.get("is_active") == "on"

        if not title:
            flash("El título es obligatorio.", "error")
            return render_template("admin/survey_form.html", survey=None)

        survey = Survey(
            title=title,
            description=description,
            is_active=is_active,
            start_date=datetime.fromisoformat(start_str) if start_str else None,
            end_date=datetime.fromisoformat(end_str) if end_str else None,
        )
        db.session.add(survey)
        db.session.commit()
        flash("Encuesta creada.", "success")
        return redirect(url_for("admin.survey_edit", survey_id=survey.id))

    return render_template("admin/survey_form.html", survey=None)


@admin_bp.route("/survey/<int:survey_id>/edit", methods=["GET", "POST"])
@operator_required
def survey_edit(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    if request.method == "POST":
        survey.title = request.form.get("title", "").strip()
        survey.description = request.form.get("description", "").strip()
        start_str = request.form.get("start_date", "").strip()
        end_str = request.form.get("end_date", "").strip()
        survey.is_active = request.form.get("is_active") == "on"
        survey.start_date = datetime.fromisoformat(start_str) if start_str else None
        survey.end_date = datetime.fromisoformat(end_str) if end_str else None
        survey.terms_text = request.form.get("terms_text", "").strip() or None
        db.session.commit()
        flash("Encuesta actualizada.", "success")

    return render_template("admin/survey_edit.html", survey=survey)


@admin_bp.route("/survey/<int:survey_id>/delete", methods=["POST"])
@admin_required
def survey_delete(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    # cascade delete
    for r in survey.responses:
        Answer.query.filter_by(response_id=r.id).delete()
    Response.query.filter_by(survey_id=survey_id).delete()
    for q in survey.questions:
        QuestionOption.query.filter_by(question_id=q.id).delete()
    Question.query.filter_by(survey_id=survey_id).delete()
    db.session.delete(survey)
    db.session.commit()
    flash("Encuesta eliminada.", "success")
    return redirect(url_for("admin.index"))


# ── Questions (API) ────────────────────────────────────────────────────────────

@admin_bp.route("/survey/<int:survey_id>/questions", methods=["GET"])
@operator_required
def questions_list(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    questions = Question.query.filter_by(
        survey_id=survey_id, parent_question_id=None
    ).order_by(Question.order).all()

    def q_to_dict(q):
        return {
            "id": q.id,
            "text": q.text,
            "question_type": q.question_type,
            "order": q.order,
            "required": q.required,
            "rating_min": q.rating_min,
            "rating_max": q.rating_max,
            "options": [
                {"id": o.id, "text": o.text, "order": o.order} for o in q.options
            ],
            "subquestions": [
                {
                    "id": sq.id,
                    "text": sq.text,
                    "question_type": sq.question_type,
                    "required": sq.required,
                    "rating_min": sq.rating_min,
                    "rating_max": sq.rating_max,
                    "parent_trigger_value": sq.parent_trigger_value,
                    "options": [
                        {"id": o.id, "text": o.text, "order": o.order}
                        for o in sq.options
                    ],
                }
                for sq in q.subquestions
            ],
        }

    return jsonify([q_to_dict(q) for q in questions])


@admin_bp.route("/survey/<int:survey_id>/questions", methods=["POST"])
@operator_required
def question_create(survey_id):
    Survey.query.get_or_404(survey_id)
    data = request.json

    # Determine next order
    max_order = (
        db.session.query(db.func.max(Question.order))
        .filter_by(survey_id=survey_id, parent_question_id=None)
        .scalar()
        or 0
    )

    q = Question(
        survey_id=survey_id,
        text=data.get("text", ""),
        question_type=data.get("question_type", "single"),
        order=max_order + 1,
        required=data.get("required", True),
        rating_min=data.get("rating_min", 1),
        rating_max=data.get("rating_max", 5),
    )
    db.session.add(q)
    db.session.flush()

    for i, opt_text in enumerate(data.get("options", [])):
        if opt_text.strip():
            db.session.add(QuestionOption(question_id=q.id, text=opt_text.strip(), order=i))

    db.session.commit()
    return jsonify({"id": q.id, "message": "Pregunta creada"}), 201


@admin_bp.route("/questions/<int:q_id>", methods=["PUT"])
@operator_required
def question_update(q_id):
    q = Question.query.get_or_404(q_id)
    data = request.json

    q.text = data.get("text", q.text)
    q.question_type = data.get("question_type", q.question_type)
    q.required = data.get("required", q.required)
    q.rating_min = data.get("rating_min", q.rating_min)
    q.rating_max = data.get("rating_max", q.rating_max)

    # Replace options
    if "options" in data:
        QuestionOption.query.filter_by(question_id=q.id).delete()
        for i, opt_text in enumerate(data["options"]):
            if opt_text.strip():
                db.session.add(
                    QuestionOption(question_id=q.id, text=opt_text.strip(), order=i)
                )

    db.session.commit()
    return jsonify({"message": "Actualizado"})


@admin_bp.route("/questions/<int:q_id>", methods=["DELETE"])
@operator_required
def question_delete(q_id):
    q = Question.query.get_or_404(q_id)
    # Delete subquestions first
    for sq in q.subquestions:
        QuestionOption.query.filter_by(question_id=sq.id).delete()
        Answer.query.filter_by(question_id=sq.id).delete()
        db.session.delete(sq)
    QuestionOption.query.filter_by(question_id=q.id).delete()
    Answer.query.filter_by(question_id=q.id).delete()
    db.session.delete(q)
    db.session.commit()
    return jsonify({"message": "Eliminado"})


@admin_bp.route("/questions/<int:q_id>/reorder", methods=["POST"])
@operator_required
def question_reorder(q_id):
    data = request.json  # {"order": [id1, id2, ...]}
    for idx, qid in enumerate(data.get("order", [])):
        Question.query.filter_by(id=qid).update({"order": idx + 1})
    db.session.commit()
    return jsonify({"message": "Reordenado"})


@admin_bp.route("/questions/<int:q_id>/subquestion", methods=["POST"])
@operator_required
def subquestion_create(q_id):
    parent = Question.query.get_or_404(q_id)
    data = request.json

    sq = Question(
        survey_id=parent.survey_id,
        text=data.get("text", ""),
        question_type=data.get("question_type", "text"),
        order=0,
        required=data.get("required", False),
        rating_min=data.get("rating_min", 1),
        rating_max=data.get("rating_max", 5),
        parent_question_id=q_id,
        parent_trigger_value=data.get("parent_trigger_value", ""),
    )
    db.session.add(sq)
    db.session.flush()

    for i, opt_text in enumerate(data.get("options", [])):
        if opt_text.strip():
            db.session.add(
                QuestionOption(question_id=sq.id, text=opt_text.strip(), order=i)
            )

    db.session.commit()
    return jsonify({"id": sq.id, "message": "Subpregunta creada"}), 201


@admin_bp.route("/subquestions/<int:sq_id>", methods=["PUT"])
@operator_required
def subquestion_update(sq_id):
    sq = Question.query.get_or_404(sq_id)
    data = request.json
    sq.text = data.get("text", sq.text)
    sq.question_type = data.get("question_type", sq.question_type)
    sq.required = data.get("required", sq.required)
    sq.rating_min = data.get("rating_min", sq.rating_min)
    sq.rating_max = data.get("rating_max", sq.rating_max)
    sq.parent_trigger_value = data.get("parent_trigger_value", sq.parent_trigger_value)

    if "options" in data:
        QuestionOption.query.filter_by(question_id=sq.id).delete()
        for i, opt_text in enumerate(data["options"]):
            if opt_text.strip():
                db.session.add(
                    QuestionOption(question_id=sq.id, text=opt_text.strip(), order=i)
                )

    db.session.commit()
    return jsonify({"message": "Actualizado"})


@admin_bp.route("/subquestions/<int:sq_id>", methods=["DELETE"])
@operator_required
def subquestion_delete(sq_id):
    sq = Question.query.get_or_404(sq_id)
    QuestionOption.query.filter_by(question_id=sq.id).delete()
    Answer.query.filter_by(question_id=sq.id).delete()
    db.session.delete(sq)
    db.session.commit()
    return jsonify({"message": "Eliminado"})


# ── Users / Import ─────────────────────────────────────────────────────────────

@admin_bp.route("/users")
@admin_required
def users_list():
    users = User.query.filter_by(is_admin=False).order_by(User.nombre).all()
    return render_template("admin/users.html", users=users)


@admin_bp.route("/users/import", methods=["GET", "POST"])
@admin_required
def users_import():
    if request.method == "POST":
        file = request.files.get("excel_file")
        if not file or not file.filename.endswith(".xlsx"):
            flash("Sube un archivo .xlsx válido.", "error")
            return redirect(request.url)

        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [
            str(cell.value).strip().lower() if cell.value else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]

        # Map column indexes flexibly
        def col(names):
            for name in names:
                for i, h in enumerate(headers):
                    if name in h:
                        return i
            return None

        # Detección automática de formato:
        # Formato A (nuevo): columnas UF, RESPONSABLE, TIPO
        # Formato B (viejo): columnas documento/residente, grupo primario
        idx_uf = col(["uf"])
        idx_responsable = col(["responsable"])
        idx_tipo = col(["tipo"])
        idx_nombre = col(["residente", "nombre", "name"])
        idx_email = col(["email", "correo", "mail"])
        idx_grupo = col(["grupo primario", "primario", "grupo_primario"])
        idx_parcela = col(["parcela"])
        idx_app = col(["app"])

        # Usar formato nuevo (UF) si existe, sino formato viejo
        usar_formato_uf = idx_uf is not None and idx_responsable is not None

        if not usar_formato_uf and idx_grupo is None:
            flash("No se encontraron columnas válidas en el Excel.", "error")
            return redirect(request.url)

        # Leer filas — un usuario por UF (formato nuevo) o por grupo primario (formato viejo)
        usuarios_vistos = {}
        total_filas = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            total_filas += 1
            if usar_formato_uf:
                uf = str(row[idx_uf]).strip()[:50] if row[idx_uf] else ""
                nombre = str(row[idx_responsable]).strip()[:200] if row[idx_responsable] else ""
                tipo = str(row[idx_tipo]).strip()[:200] if (idx_tipo is not None and row[idx_tipo]) else ""
                email = str(row[idx_email]).strip()[:200] if (idx_email is not None and row[idx_email]) else ""
                if uf and nombre:
                    usuarios_vistos[uf] = {
                        "documento": uf,
                        "nombre": nombre,
                        "email": email,
                        "grupo_primario": tipo,
                        "parcela": uf,
                        "app_access": True,
                    }
            else:
                grupo = str(row[idx_grupo]).strip()[:200] if (idx_grupo is not None and row[idx_grupo]) else ""
                nombre = str(row[idx_nombre]).strip()[:200] if row[idx_nombre] else ""
                email = str(row[idx_email]).strip()[:200] if (idx_email is not None and row[idx_email]) else ""
                parcela_val = str(row[idx_parcela]).strip()[:100] if (idx_parcela is not None and row[idx_parcela]) else ""
                app_val = str(row[idx_app]).strip().lower() if (idx_app is not None and row[idx_app]) else "si"
                if grupo and nombre and grupo not in usuarios_vistos:
                    usuarios_vistos[grupo] = {
                        "documento": grupo,
                        "nombre": grupo,
                        "email": email,
                        "grupo_primario": grupo,
                        "parcela": parcela_val,
                        "app_access": app_val in ("si", "sí", "yes", "true", "1"),
                    }

        rows_data = list(usuarios_vistos.values())
        skipped = total_filas - len(rows_data)

        # Una sola consulta para obtener grupos ya existentes
        grupos_en_excel = [r["documento"] for r in rows_data]
        existentes = {
            u.documento: u
            for u in User.query.filter(User.documento.in_(grupos_en_excel)).all()
        }

        created = 0
        updated = 0
        nuevos = {}

        for r in rows_data:
            if r["documento"] in existentes:
                u = existentes[r["documento"]]
                u.nombre = r["nombre"]
                u.email = r["email"]
                u.grupo_primario = r["grupo_primario"]
                u.parcela = r["parcela"]
                u.app_access = r["app_access"]
                updated += 1
            else:
                nuevos[r["documento"]] = r

        nuevos = list(nuevos.values())

        # Insertar nuevos — contraseña inicial = grupo primario
        from werkzeug.security import generate_password_hash as gph
        def fast_hash(password):
            return gph(password, method="pbkdf2:sha256", salt_length=8)

        BATCH = 500
        for i in range(0, len(nuevos), BATCH):
            lote = nuevos[i:i + BATCH]
            db.session.bulk_insert_mappings(
                User,
                [
                    {
                        "documento": r["documento"],
                        "nombre": r["nombre"],
                        "email": r["email"],
                        "grupo_primario": r["grupo_primario"],
                        "parcela": r["parcela"],
                        "app_access": r["app_access"],
                        "is_admin": False,
                        "password_hash": fast_hash(r["documento"]),
                        "created_at": datetime.utcnow(),
                    }
                    for r in lote
                ],
            )
            created += len(lote)

        db.session.commit()
        flash(
            f"Importación completada: {created} grupos creados, {updated} actualizados, {skipped} filas duplicadas omitidas.",
            "success",
        )
        return redirect(url_for("admin.users_list"))

    return render_template("admin/users_import.html")


@admin_bp.route("/users/<int:user_id>/toggle-operator", methods=["POST"])
@admin_required
def user_toggle_operator(user_id):
    user = User.query.get_or_404(user_id)
    if not user.is_admin:
        user.is_operator = not user.is_operator
        db.session.commit()
        rol = "operador" if user.is_operator else "usuario"
        flash(f"{user.nombre} ahora es {rol}.", "success")
    return redirect(url_for("admin.users_list"))


@admin_bp.route("/users/clear", methods=["POST"])
@admin_required
def users_clear():
    """Elimina todos los usuarios no-admin y sus respuestas."""
    users = User.query.filter_by(is_admin=False).all()
    for u in users:
        for r in u.responses:
            Answer.query.filter_by(response_id=r.id).delete()
        Response.query.filter_by(user_id=u.id).delete()
    User.query.filter_by(is_admin=False).delete()
    db.session.commit()
    flash("Todos los usuarios (no admin) fueron eliminados.", "success")
    return redirect(url_for("admin.users_list"))


@admin_bp.route("/users/<int:user_id>/reset-password", methods=["POST"])
@admin_required
def user_reset_password(user_id):
    user = User.query.get_or_404(user_id)
    user.set_password(user.documento)
    db.session.commit()
    return jsonify({"message": f"Contraseña de {user.nombre} restablecida al documento."})


@admin_bp.route("/users/<int:user_id>/delete", methods=["POST"])
@admin_required
def user_delete(user_id):
    user = User.query.get_or_404(user_id)
    for r in user.responses:
        Answer.query.filter_by(response_id=r.id).delete()
    Response.query.filter_by(user_id=user_id).delete()
    db.session.delete(user)
    db.session.commit()
    flash(f"Usuario {user.nombre} eliminado.", "success")
    return redirect(url_for("admin.users_list"))


# ── Results Dashboard ──────────────────────────────────────────────────────────

@admin_bp.route("/survey/<int:survey_id>/results")
@admin_required
def survey_results(survey_id):
    survey = Survey.query.get_or_404(survey_id)
    total_users = User.query.filter_by(is_admin=False).count()
    total_responses = Response.query.filter_by(survey_id=survey_id, is_complete=True).count()
    pending_count = total_users - total_responses

    questions = Question.query.filter_by(
        survey_id=survey_id, parent_question_id=None
    ).order_by(Question.order).all()

    results = []
    for q in questions:
        q_data = _build_question_result(q)
        q_data["subquestions"] = [_build_question_result(sq) for sq in q.subquestions]
        results.append(q_data)

    # Participación por parcela
    group_stats = (
        db.session.query(User.parcela, db.func.count(User.id))
        .filter(User.is_admin == False)
        .group_by(User.parcela)
        .order_by(User.parcela)
        .all()
    )
    responded_map = dict(
        db.session.query(User.parcela, db.func.count(Response.id))
        .join(Response, Response.user_id == User.id)
        .filter(Response.survey_id == survey_id, Response.is_complete == True)
        .group_by(User.parcela)
        .all()
    )
    group_participation = sorted(
        [
            {
                "group": g or "Sin grupo",
                "total": t,
                "responded": responded_map.get(g, 0),
                "pct": round(responded_map.get(g, 0) / t * 100, 1) if t else 0,
            }
            for g, t in group_stats
        ],
        key=lambda x: x["pct"],
        reverse=True,
    )

    return render_template(
        "admin/survey_results.html",
        survey=survey,
        results=results,
        total_users=total_users,
        total_responses=total_responses,
        pending_count=pending_count,
        group_participation=group_participation,
    )


def _build_question_result(q):
    from collections import Counter
    answers = Answer.query.filter_by(question_id=q.id).all()
    total_answers = len(answers)

    if q.question_type in ("single", "multiple"):
        option_counts = {o.id: {"text": o.text, "count": 0} for o in q.options}
        for a in answers:
            for oid in a.get_option_ids():
                if oid in option_counts:
                    option_counts[oid]["count"] += 1
        chart_data = list(option_counts.values())
        # Elegir tipo de gráfico: torta si ≤5 opciones y es single, barras si más
        chart_type = "pie" if (q.question_type == "single" and len(q.options) <= 5) else "bar"
        return {
            "id": q.id, "text": q.text, "type": q.question_type,
            "total_answers": total_answers,
            "chart_data": chart_data,
            "chart_type": chart_type,
        }

    elif q.question_type == "rating":
        rating_counts = Counter()
        for a in answers:
            if a.text_value:
                try:
                    rating_counts[int(a.text_value)] += 1
                except ValueError:
                    pass
        avg = sum(k * v for k, v in rating_counts.items()) / total_answers if total_answers else 0
        chart_data = [{"text": str(v), "count": cnt} for v, cnt in sorted(rating_counts.items())]
        return {
            "id": q.id, "text": q.text, "type": q.question_type,
            "total_answers": total_answers,
            "chart_data": chart_data,
            "chart_type": "bar",
            "average": round(avg, 2),
        }

    else:  # text
        return {
            "id": q.id, "text": q.text, "type": q.question_type,
            "total_answers": total_answers,
            "chart_data": [{"text": a.text_value} for a in answers if a.text_value],
            "chart_type": "text",
        }


# ── Export Excel ───────────────────────────────────────────────────────────────

@admin_bp.route("/survey/<int:survey_id>/export")
@admin_required
def survey_export(survey_id):
    import io
    from collections import Counter
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    from flask import send_file

    survey = Survey.query.get_or_404(survey_id)
    wb = Workbook()

    BLUE   = "2563EB"
    LBLUE  = "DBEAFE"
    GREEN  = "16A34A"
    LGREEN = "DCFCE7"
    GRAY   = "F8FAFC"
    DGRAY  = "64748B"

    def header_style(cell, bg=BLUE):
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def title_style(cell):
        cell.font = Font(bold=True, size=13, color=BLUE)

    def thin_border():
        s = Side(style="thin", color="E2E8F0")
        return Border(left=s, right=s, top=s, bottom=s)

    # ── Hoja 1: Resumen ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"Resultados: {survey.title}"
    title_style(c)
    c.alignment = Alignment(horizontal="left", vertical="center")

    ws["A2"].value = survey.description or ""
    ws["A2"].font = Font(color=DGRAY, italic=True)

    total_users = User.query.filter_by(is_admin=False).count()
    total_resp  = Response.query.filter_by(survey_id=survey_id, is_complete=True).count()

    ws["A4"].value = "Total usuarios"
    ws["B4"].value = total_users
    ws["A5"].value = "Respondieron"
    ws["B5"].value = total_resp
    ws["A6"].value = "Pendientes"
    ws["B6"].value = total_users - total_resp
    ws["A7"].value = "Participación"
    ws["B7"].value = f"{round(total_resp/total_users*100,1) if total_users else 0}%"
    for r in range(4, 8):
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"].alignment = Alignment(horizontal="center")

    # Participación por sector
    ws["A9"].value = "Participación por sector"
    ws["A9"].font = Font(bold=True, size=11, color=BLUE)

    headers = ["Sector", "Total", "Respondieron", "Pendientes", "%"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=10, column=i, value=h)
        header_style(c)

    group_stats = (
        db.session.query(User.grupo_primario, db.func.count(User.id))
        .filter(User.is_admin == False)
        .group_by(User.grupo_primario)
        .order_by(User.grupo_primario)
        .all()
    )
    responded_map = dict(
        db.session.query(User.grupo_primario, db.func.count(Response.id))
        .join(Response, Response.user_id == User.id)
        .filter(Response.survey_id == survey_id, Response.is_complete == True)
        .group_by(User.grupo_primario)
        .all()
    )
    data_start = 11
    for i, (g, t) in enumerate(group_stats):
        r = data_start + i
        resp = responded_map.get(g, 0)
        pct  = round(resp / t * 100, 1) if t else 0
        vals = [g or "Sin grupo", t, resp, t - resp, f"{pct}%"]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = thin_border()
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=GRAY)

    # Torta de participación
    pie = PieChart()
    pie.title = "Respondieron por sector"
    pie.style = 10
    pie.width = 14
    pie.height = 10
    labels_ref = Reference(ws, min_col=1, min_row=data_start, max_row=data_start + len(group_stats) - 1)
    data_ref   = Reference(ws, min_col=3, min_row=data_start, max_row=data_start + len(group_stats) - 1)
    pie.add_data(data_ref)
    pie.set_categories(labels_ref)
    pie.dataLabels = None
    ws.add_chart(pie, f"G9")

    # ── Hojas por pregunta ─────────────────────────────────────────────────────
    questions = Question.query.filter_by(
        survey_id=survey_id, parent_question_id=None
    ).order_by(Question.order).all()

    for qi, q in enumerate(questions, 1):
        _add_question_sheet(wb, q, qi, thin_border, header_style, title_style,
                            BLUE, LBLUE, GREEN, GRAY, DGRAY)
        for sqi, sq in enumerate(q.subquestions, 1):
            _add_question_sheet(wb, sq, qi, thin_border, header_style, title_style,
                                BLUE, LBLUE, GREEN, GRAY, DGRAY,
                                label=f"P{qi}.{sqi}")

    # ── Hoja: Respuestas texto libre ───────────────────────────────────────────
    text_qs = [q for q in questions if q.question_type == "text"]
    text_qs += [sq for q in questions for sq in q.subquestions if sq.question_type == "text"]
    if text_qs:
        wst = wb.create_sheet("Texto libre")
        wst.column_dimensions["A"].width = 40
        wst.column_dimensions["B"].width = 55
        row = 1
        for q in text_qs:
            wst.cell(row=row, column=1, value=q.text).font = Font(bold=True, color=BLUE)
            row += 1
            answers = Answer.query.filter_by(question_id=q.id).all()
            for a in answers:
                if a.text_value:
                    wst.cell(row=row, column=2, value=a.text_value).alignment = Alignment(wrap_text=True)
                    row += 1
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"resultados_{survey.title[:30].replace(' ','_')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def _add_question_sheet(wb, q, qi, thin_border, header_style, title_style,
                         BLUE, LBLUE, GREEN, GRAY, DGRAY, label=None):
    from collections import Counter
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    sheet_name = f"{'P' if not label else label}{qi}"[:31]
    # Avoid duplicate sheet names
    existing = [s.title for s in wb.worksheets]
    base = sheet_name
    idx = 2
    while sheet_name in existing:
        sheet_name = f"{base}_{idx}"
        idx += 1

    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = q.text
    title_style(c)
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 40

    answers = Answer.query.filter_by(question_id=q.id).all()
    total = len(answers)
    ws["A2"].value = f"{total} respuesta(s)"
    ws["A2"].font = Font(color=DGRAY, size=10)

    if q.question_type in ("single", "multiple"):
        option_counts = {o.id: {"text": o.text, "count": 0} for o in q.options}
        for a in answers:
            for oid in a.get_option_ids():
                if oid in option_counts:
                    option_counts[oid]["count"] += 1

        # Headers
        for col, h in enumerate(["Opción", "Respuestas", "%"], 1):
            c = ws.cell(row=4, column=col, value=h)
            header_style(c)

        data_rows = list(option_counts.values())
        for i, d in enumerate(data_rows):
            r = 5 + i
            pct = round(d["count"] / total * 100, 1) if total else 0
            ws.cell(row=r, column=1, value=d["text"]).border = thin_border()
            ws.cell(row=r, column=2, value=d["count"]).border = thin_border()
            ws.cell(row=r, column=3, value=f"{pct}%").border = thin_border()
            if i % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor="F8FAFC")

        # Chart
        chart_row_start = 5
        chart_row_end   = 4 + len(data_rows)
        use_pie = (q.question_type == "single" and len(data_rows) <= 5)

        if use_pie:
            chart = PieChart()
            chart.style = 10
        else:
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.grouping = "clustered"

        chart.title = q.text[:40]
        chart.width  = 14
        chart.height = 10
        labels_ref = Reference(ws, min_col=1, min_row=chart_row_start, max_row=chart_row_end)
        data_ref   = Reference(ws, min_col=2, min_row=chart_row_start - 1, max_row=chart_row_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        ws.add_chart(chart, "E4")

    elif q.question_type == "rating":
        rating_counts = Counter()
        for a in answers:
            if a.text_value:
                try:
                    rating_counts[int(a.text_value)] += 1
                except ValueError:
                    pass
        avg = sum(k * v for k, v in rating_counts.items()) / total if total else 0
        ws["A2"].value = f"{total} respuesta(s)  |  Promedio: {round(avg,2)}"

        for col, h in enumerate(["Valor", "Cantidad", "%"], 1):
            c = ws.cell(row=4, column=col, value=h)
            header_style(c)

        sorted_items = sorted(rating_counts.items())
        for i, (val, cnt) in enumerate(sorted_items):
            r = 5 + i
            ws.cell(row=r, column=1, value=val).border = thin_border()
            ws.cell(row=r, column=2, value=cnt).border = thin_border()
            ws.cell(row=r, column=3, value=f"{round(cnt/total*100,1) if total else 0}%").border = thin_border()

        chart = BarChart()
        chart.type   = "col"
        chart.style  = 10
        chart.title  = q.text[:40]
        chart.width  = 14
        chart.height = 10
        data_ref   = Reference(ws, min_col=2, min_row=4, max_row=4 + len(sorted_items))
        labels_ref = Reference(ws, min_col=1, min_row=5, max_row=4 + len(sorted_items))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(labels_ref)
        ws.add_chart(chart, "E4")
